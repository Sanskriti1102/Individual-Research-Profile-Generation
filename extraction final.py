import openpyxl
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By
import time

# Function to fetch scholar ID and email from input file
def fetch_scholar_id(professor_name):
    input_wb = openpyxl.load_workbook('input.xlsx')
    input_sheet = input_wb.active

    scholar_id = None
    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        name, scholar_id, scholar_email = row
        if name.lower() == professor_name.lower():
            break

    input_wb.close()
    return scholar_id

# Function to get Google Scholar data and store it in the Excel sheet
def get_google_scholar_data(professor_name, scholar_id):
    url = f'https://scholar.google.com/citations?hl=en&user={scholar_id}&view_op=list_works&sortby=pubdate'
    
    browser_options = ChromeOptions()
    browser_options.add_argument('--headless')
    driver = Chrome(options=browser_options)
    driver.maximize_window()
    driver.get(url)

    count = 0
    num_of_literatures = driver.find_elements(By.XPATH, "//table[@id='gsc_a_t']//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr']")
    for _ in num_of_literatures:
        count += 1

    while True:
        n_count = 0
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)
        try:
            show_more_button = driver.find_element(By.XPATH, "//button[@id='gsc_bpf_more']").click()
        except Exception as e:
            break
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(5)
        num_of_literatures = driver.find_elements(By.XPATH, "//table[@id='gsc_a_t']//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr']")
        for _ in num_of_literatures:
            n_count += 1
        if n_count == count:
            break
        else:
            count = n_count

    fetched_data = False
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ['Title', 'Authors', 'Publication date', 'Journal', 'Volume', 'Issue', 'Pages',
               'Publisher', 'Total citations', 'Document Type', 'DOI', 'Link']
    sheet.append(headers)

    for i in range(1, count + 1):
        try:
            publications_url = driver.find_element(By.XPATH, "//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr'][" + str(i) + "]//td[@class='gsc_a_t']//a").get_attribute('href')
            title = driver.find_element(By.XPATH, "//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr'][" + str(i) + "]//td[@class='gsc_a_t']//a").text.strip()
            authors = driver.find_element(By.XPATH, "//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr'][" + str(i) + "]//td[@class='gsc_a_t']//div[@class='gs_gray']").text.strip()
            publication_date = driver.find_element(By.XPATH, "//tbody[@id='gsc_a_b']//tr[@class='gsc_a_tr'][" + str(i) + "]//td[@class='gsc_a_y']").text.strip()
            doc_type = "Other"  # Modify this based on your document types
            sheet.append([title, authors, publication_date, '', '', '', '', '', '', doc_type, '', publications_url])
            fetched_data = True
        except Exception as e:
            print(f"Error occurred: {e}")

    driver.quit()

    if fetched_data:
        file_name = f'{professor_name}.xlsx'
        workbook.save(file_name)
        print(f'All data saved to Excel file: {file_name}')
    else:
        print(f'No literature data fetched for {professor_name}.')

# Main function
def main():
    while True:
        professor_name_input = input("Enter the professor's name: ")

        scholar_id = fetch_scholar_id(professor_name_input)

        if scholar_id is None:
            print(f"Professor '{professor_name_input}' not found in the input sheet.")
            continue

        get_google_scholar_data(professor_name_input, scholar_id)

        continue_option = input("Choose an option:\n1. Extract for another professor\n2. Exit\nOption: ")
        if continue_option == '2':
            print("Exit the program.")
            break

# Call the main function
main()
