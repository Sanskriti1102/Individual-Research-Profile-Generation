import openpyxl
from wordcloud import WordCloud
import matplotlib.pyplot as plt

# Function to fetch literature data from an Excel sheet
def get_literature_data(sheet):
    # Assuming the Excel sheet has headers in the first row
    headers = [cell.value for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def filter_and_save_data(keyword_input, data, professor_name):
    filtered_data = [entry for entry in data if 
                     keyword_input.lower() in entry.get('Title', '').lower() or 
                     keyword_input.lower() in entry.get('Authors', '').lower()]
    print("\n")
    print(f'\nSearching for keyword: {keyword_input}\n')
    print(f'\nFound {len(filtered_data)} matching entries.\n')

    if filtered_data:
        wb_filtered = openpyxl.Workbook()
        sheet_filtered = wb_filtered.active

        headers = list(data[0].keys())
        sheet_filtered.append(headers)

        for entry in filtered_data:
            row = [entry[header] for header in headers]
            sheet_filtered.append(row)

        file_name = f'filtered_data_{professor_name}_{keyword_input}.xlsx'
        wb_filtered.save(file_name)
        print(f'\nData saved to Excel file: {file_name}\n')
    else:
        print('No data matching the keyword found.')



def generate_word_cloud(data, professor_name):
    # Get all titles and authors
    titles = [entry['Title'] for entry in data]
    authors = []

    for entry in data:
        # Check if the entry has authors
        if 'Authors' in entry and entry['Authors'] is not None:
            # Extract author names from the "Author" column
            author_names = entry['Authors'].split(', ')
            authors.extend(author_names)

    # Generate word cloud for titles
    titles_text = ' '.join(titles)
    title_wordcloud = WordCloud(width=400, height=400, background_color='white',
                                colormap='viridis',  # Change the color map
                                contour_color='red',  # Add a contour color
                                contour_width=2,  # Set contour width
                                ).generate(titles_text)

    # Generate word cloud for authors
    if authors:
        authors_text = ' '.join(authors)
        author_wordcloud = WordCloud(width=400, height=400, background_color='white',
                                     colormap='plasma',  # Change the color map
                                     contour_color='yellow',  # Add a contour color
                                     contour_width=2,  # Set contour width
                                     ).generate(authors_text)
    else:
        author_wordcloud = None

    # Display word clouds
    fig, ax = plt.subplots(1, 2, figsize=(10, 10))

    # Define font properties
    font = {'family': 'serif',
        'color':  'darkred',
        'weight': 'bold',
        'size': 10,
        }

    # Plot title word cloud
    ax[0].imshow(title_wordcloud, interpolation='bilinear')
    ax[0].set_title(f'Title Word Cloud for {professor_name}', fontsize=10, fontweight='bold', color='black',fontdict=font)  # Style the title
    ax[0].axis('off')

    # Plot author word cloud if available
    if author_wordcloud:
        ax[1].imshow(author_wordcloud, interpolation='bilinear')
        ax[1].set_title(f'Author Word Cloud for {professor_name}', fontsize=10, fontweight='bold', color='green',fontdict=font)  # Style the title
        ax[1].axis('off')
    else:
        ax[1].axis('off')
        ax[1].set_title(f'No Authors Available', fontsize=16, fontweight='bold', color='red')  # Style the title

    # Adjust subplot arrangement to center the word clouds
    plt.subplots_adjust(left=0.2, right=0.8, top=0.8, bottom=0.2, wspace=0.3, hspace=0.3)

    # Show the plot
    plt.tight_layout()  # Adjust layout to prevent overlap
    plt.show()




# Main function
def main():
    professor_name_input = None

    while True:
        if professor_name_input is None:
            professor_name_input = input("\nEnter the professor's name: ")

        try:
            wb = openpyxl.load_workbook(f'{professor_name_input}.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            print("\n")
            print(f"File '{professor_name_input}.xlsx' not found.")
            print("\n")
            professor_name_input = None
            continue

        data = get_literature_data(sheet)

        while True:
            print("\nChoose an option:")
            print("1. Filter papers by keyword")
            print("2. Retrieve all papers")
            print("3. Exit")
            sub_option = input("Option: ")

            if sub_option == '1':
                generate_word_cloud(data, professor_name_input)
                # keyword_input = input("Enter a keyword to filter papers: ")
                # filter_and_save_data(keyword_input, data, professor_name_input)
                
                # Filter and recommend titles
                keyword_input = input("\n Enter a keyword to filter papers: ")
                filtered_data = [entry for entry in data if 
                                 keyword_input.lower() in entry.get('Title', '').lower() or 
                                 keyword_input.lower() in entry.get('Authors', '').lower()]
                if filtered_data:
                    
                    print("\nRecommended titles:\n")
                    # print("\n")
                    for entry in filtered_data:
                        print(entry['Title'])
                    filter_and_save_data(keyword_input, data, professor_name_input)
                else:
                    print("\n No matching entries found.\n")
                    
            elif sub_option == '2':
                file_name = f'{professor_name_input}.xlsx'
                print("\nTitles:\n")
                for entry in data:
                    print(entry['Title'])
                wb.save(file_name)
                print("\n")
                print(f'All data saved to Excel file: {file_name}')
                print("\n")
            elif sub_option == '3':
                print("\nExiting the program.")
                return
            else:
                print("\nInvalid option selected.")
                continue
            
            # continue_option = input("Do you want to continue? (yes/no) or (y/n) or (Y/N) or (Yes/No) or (YES/NO): ")
            # if continue_option.lower() not in ('yes', 'y','Y','Yes','YES'):
            #     break

            print("\nChoose options:")
            print("1. Continue for the same Professor")
            print("2. Get the results for another professor")
            print("3. Exit")
            option = input("Option: ")

            if option == '1':
                continue
            elif option == '2':
                professor_name_input = input("\nEnter the professor's name: ")
                break
            elif option == '3':
                print("\nExiting the program.")
                return
            else:
                print("\nInvalid option selected.")

# Call the main function
main()
