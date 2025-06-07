import openpyxl
import logging
from selenium.webdriver import Chrome, ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class GoogleScholarScraper:
    def __init__(self, professor_name, scholar_id):
        self.professor_name = professor_name
        self.scholar_id = scholar_id
        self.driver = None
        self.base_url = f'https://scholar.google.com/citations?hl=en&user={self.scholar_id}&view_op=list_works&sortby=pubdate'
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.wait = None
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    def setup_driver(self):
        options = ChromeOptions()
        options.add_argument('--headless')
        self.driver = Chrome(options=options)
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 10)

    def teardown_driver(self):
        if self.driver:
            self.driver.quit()

    def scroll_to_load_all(self):
        count = 0
        while True:
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            try:
                self.wait.until(EC.presence_of_element_located((By.ID, 'gsc_bpf_more')))
                self.driver.find_element(By.ID, "gsc_bpf_more").click()
                self.wait.until(EC.staleness_of(self.driver.find_element(By.ID, "gsc_bpf_more")))
            except:
                break
            new_count = len(self.driver.find_elements(By.CSS_SELECTOR, "tr.gsc_a_tr"))
            if new_count == count:
                break
            count = new_count

    def scrape_publications(self):
        # Clean header without Journal, Volume, Issue, Pages, Publisher, Total citations, DOI
        self.sheet.append(['Title', 'Authors', 'Publication date', 'Document Type', 'Link'])

        rows = self.driver.find_elements(By.CSS_SELECTOR, "tr.gsc_a_tr")

        for idx, row in enumerate(rows, 1):
            try:
                title_el = row.find_element(By.CSS_SELECTOR, "td.gsc_a_t a")
                title = title_el.text.strip()
                link = title_el.get_attribute("href")

                author_journal = row.find_elements(By.CSS_SELECTOR, "td.gsc_a_t div.gs_gray")
                authors = author_journal[0].text.strip() if len(author_journal) > 0 else ""

                pub_date = row.find_element(By.CSS_SELECTOR, "td.gsc_a_y span.gsc_a_h.gsc_a_hc.gs_ibl").text.strip()

                self.sheet.append([title, authors, pub_date, 'Other', link])
            except Exception as e:
                logging.error(f"Error scraping publication {idx}: {e}")

    def save_to_excel(self):
        file_name = f'{self.professor_name}.xlsx'
        self.workbook.save(file_name)
        logging.info(f"Data saved to {file_name}")

    def run(self):
        logging.info(f"Starting scraping for {self.professor_name}")
        self.setup_driver()
        self.driver.get(self.base_url)
        self.scroll_to_load_all()
        self.scrape_publications()
        self.teardown_driver()
        self.save_to_excel()


def fetch_scholar_id(professor_name):
    wb = openpyxl.load_workbook('D:/practice external/projects videos/Individual-Research-Profile-Generation/input.xlsx')
    sheet = wb.active
    scholar_id = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, s_id, _ = row
        if name and name.strip().lower() == professor_name.strip().lower():
            scholar_id = s_id
            break
    wb.close()
    return scholar_id


def main():
    while True:
        professor_name = input("Enter the professor's name: ")
        scholar_id = fetch_scholar_id(professor_name)

        if not scholar_id:
            logging.warning(f"Professor '{professor_name}' not found in input sheet.")
            continue

        scraper = GoogleScholarScraper(professor_name, scholar_id)
        scraper.run()

        choice = input("Do you want to scrape another professor? (yes/no): ").lower()
        if choice != 'yes':
            break


if __name__ == '__main__':
    main()
